from django.shortcuts import render
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import io
import logging
from datetime import datetime
from django.utils import timezone
from .models import Customer
from .serializers import CustomerSerializer

# Configure logger for customer operations
logger = logging.getLogger('customers')

class CustomerPagination(PageNumberPagination):
    """
    Pagination class for customers list.
    """
    page_size = 15
    page_size_query_param = 'page_size'
    max_page_size = 100

class CustomerViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing customers.
    Provides CRUD operations for customer management.
    """
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    pagination_class = CustomerPagination
    
    def get_queryset(self):
        """
        All authenticated users can access all customers.
        Optionally filter based on query parameters.
        """
        # Start with all customers
        queryset = Customer.objects.all()
        
        # Apply archived filter - toggle between active and archived views
        show_archived = self.request.query_params.get('show_archived', 'false').lower() == 'true'
        if show_archived:
            # Show ONLY archived customers
            queryset = queryset.filter(is_archived=True)
            logger.info(f"Showing archived customers only - Found {queryset.count()} archived customers")
        else:
            # Show ONLY active (non-archived) customers
            queryset = queryset.filter(is_archived=False)
            logger.info(f"Showing active customers only - Found {queryset.count()} active customers")
        
        # Apply search filter if provided
        search = self.request.query_params.get('search', None)
        
        if search:
            queryset = queryset.filter(
                name__icontains=search
            ) | queryset.filter(
                brand__icontains=search
            ) | queryset.filter(
                sales_channel__icontains=search
            )
            logger.info(f"Filtering customers by search: '{search}' - Found {queryset.count()} customers")
        
        return queryset.order_by('name')
    
    def perform_create(self, serializer):
        """
        Set the added_by field to the current user when creating.
        Set to None if user is not authenticated.
        """
        user = self.request.user if self.request.user.is_authenticated else None
        serializer.save(added_by=user)
    
    def destroy(self, request, *args, **kwargs):
        """Archive a customer instead of deleting"""
        instance = self.get_object()
        instance.is_archived = True
        instance.date_archived = timezone.now()
        instance.archived_by = request.user if request.user.is_authenticated else None
        instance.save(update_fields=['is_archived', 'date_archived', 'archived_by'])
        
        return Response({
            "message": "Customer archived successfully",
            "customer": CustomerSerializer(instance).data
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def search(self, request):
        """
        Custom search endpoint for customers.
        """
        return self.list(request)

    @action(detail=False, methods=['get'])
    def list_all(self, request):
        """
        Lightweight endpoint for dropdown use - returns id, name, and location only.
        No pagination for efficient single-request loading.
        """
        queryset = Customer.objects.filter(is_archived=False).values('id', 'name', 'brand', 'sales_channel').order_by('name')
        return Response(list(queryset))


@method_decorator(csrf_exempt, name='dispatch')
class CustomerExportView(APIView):
    """Export customers to PDF or Excel format"""
    
    # Available columns for export
    AVAILABLE_COLUMNS = {
        'id': 'ID',
        'name': 'Name',
        'brand': 'Brand',
        'sales_channel': 'Sales Channel',
        'date_added': 'Date Added',
        'added_by_name': 'Added By',
    }
    
    def _get_cell_value(self, customer, column):
        """Get cell value for export"""
        if column == 'id':
            return customer.id
        elif column == 'name':
            return customer.name
        elif column == 'brand':
            return customer.brand or ''
        elif column == 'sales_channel':
            return customer.sales_channel or ''
        elif column == 'date_added':
            return customer.date_added.strftime('%Y-%m-%d') if customer.date_added else ''
        elif column == 'added_by_name':
            if customer.added_by:
                if hasattr(customer.added_by, 'profile') and customer.added_by.profile:
                    return customer.added_by.profile.full_name or customer.added_by.username
                return customer.added_by.username
            return ''
        return ''
    
    def _sort_customers(self, customers, sort_field, sort_direction):
        """Sort customers by field and direction"""
        reverse = sort_direction == 'desc'
        
        if sort_field == 'date_added':
            return sorted(customers, key=lambda c: c.date_added or datetime.min.date(), reverse=reverse)
        elif sort_field == 'added_by_name':
            return sorted(customers, key=lambda c: 
                (c.added_by.profile.full_name if c.added_by and hasattr(c.added_by, 'profile') and c.added_by.profile 
                 else c.added_by.username if c.added_by else ''), reverse=reverse)
        elif sort_field in ['id']:
            return sorted(customers, key=lambda c: getattr(c, sort_field, 0), reverse=reverse)
        elif sort_field in ['name', 'brand', 'sales_channel']:
            return sorted(customers, key=lambda c: getattr(c, sort_field, '').lower(), reverse=reverse)
        return customers
    
    def post(self, request):
        """Export customers based on provided options"""
        # Parse request data
        columns = request.data.get('columns', list(self.AVAILABLE_COLUMNS.keys()))
        sort_field = request.data.get('sort_field', 'id')
        sort_direction = request.data.get('sort_direction', 'asc')
        export_format = request.data.get('format', 'excel')
        
        # Validate columns
        valid_columns = [c for c in columns if c in self.AVAILABLE_COLUMNS]
        if not valid_columns:
            return Response({
                "error": "No valid columns specified"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get customers (exclude archived)
        customers = list(Customer.objects.select_related('added_by__profile').filter(is_archived=False))
        
        # Sort customers
        customers = self._sort_customers(customers, sort_field, sort_direction)
        
        # Generate export
        if export_format == 'pdf':
            return self._generate_pdf(customers, valid_columns)
        else:
            return self._generate_excel(customers, valid_columns)
    
    def _generate_excel(self, customers, columns):
        """Generate Excel file"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Customers"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            headers = [self.AVAILABLE_COLUMNS[col] for col in columns]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data rows
            for row_idx, customer in enumerate(customers, 2):
                for col_idx, column in enumerate(columns, 1):
                    value = self._get_cell_value(customer, column)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
            
            # Adjust column widths
            column_widths = {
                'id': 8,
                'name': 25,
                'contact_email': 30,
                'phone': 15,
                'location': 25,
                'date_added': 15,
                'added_by_name': 20,
            }
            for col_idx, column in enumerate(columns, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = column_widths.get(column, 15)
            
            # Create response
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            timestamp = datetime.now().strftime('%Y-%m-%d')
            filename = f"customers_export_{timestamp}.xlsx"
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except ImportError:
            return Response({
                "error": "Excel export not available. Please install openpyxl."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _generate_pdf(self, customers, columns):
        """Generate PDF file"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            output = io.BytesIO()
            
            # Use landscape for many columns
            page_size = landscape(letter) if len(columns) > 5 else letter
            doc = SimpleDocTemplate(output, pagesize=page_size, topMargin=0.5*inch, bottomMargin=0.5*inch)
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                alignment=1,  # Center
                spaceAfter=12
            )
            elements.append(Paragraph("Customers Export", title_style))
            
            # Subtitle with date and count
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1,
                textColor=colors.grey,
                spaceAfter=20
            )
            elements.append(Paragraph(
                f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Total Records: {len(customers)}",
                subtitle_style
            ))
            
            # Table data
            headers = [self.AVAILABLE_COLUMNS[col] for col in columns]
            table_data = [headers]
            
            for customer in customers:
                row = [str(self._get_cell_value(customer, col)) for col in columns]
                table_data.append(row)
            
            # Create table
            table = Table(table_data, repeatRows=1)
            
            # Table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ])
            table.setStyle(table_style)
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            output.seek(0)
            
            timestamp = datetime.now().strftime('%Y-%m-%d')
            filename = f"customers_export_{timestamp}.pdf"
            
            response = HttpResponse(output.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except ImportError:
            return Response({
                "error": "PDF export not available. Please install reportlab."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@method_decorator(csrf_exempt, name='dispatch')
class UnarchiveCustomerView(APIView):
    """Unarchive (restore) an archived customer"""
    
    def post(self, request, pk=None):
        """Restore an archived customer"""
        try:
            customer = Customer.objects.get(pk=pk)
        except Customer.DoesNotExist:
            return Response({"error": "Customer not found"}, status=status.HTTP_404_NOT_FOUND)
        
        if not customer.is_archived:
            return Response({"error": "Customer is not archived"}, status=status.HTTP_400_BAD_REQUEST)
        
        customer.is_archived = False
        customer.date_archived = None
        customer.archived_by = None
        customer.save(update_fields=['is_archived', 'date_archived', 'archived_by'])
        
        return Response({
            "message": "Customer unarchived successfully",
            "customer": CustomerSerializer(customer).data
        }, status=status.HTTP_200_OK)
