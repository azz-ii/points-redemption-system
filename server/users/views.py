from django.shortcuts import render
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
import logging
import io
from datetime import datetime
from utils.email_service import send_account_created_email, send_password_reset_link_email, send_password_changed_email
from utils.validators import validate_password_strength
from points_audit.utils import log_points_change, bulk_log_points_changes, generate_batch_id
from points_audit.models import PointsAuditLog

# Configure logger for user operations
logger = logging.getLogger('email')

from .models import UserProfile
from .serializers import UserSerializer, UserListSerializer, SalesAgentOptionSerializer


class UserPagination(PageNumberPagination):
    """
    Pagination class for users list.
    """
    page_size = 15
    page_size_query_param = 'page_size'
    max_page_size = 100


class UserViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing users.
    Provides CRUD operations for user management.
    """
    queryset = User.objects.filter(is_superuser=False).select_related('profile')
    serializer_class = UserListSerializer
    pagination_class = UserPagination
    
    def get_queryset(self):
        """
        All authenticated users can access all users.
        Optionally filter based on query parameters.
        """
        queryset = User.objects.filter(is_superuser=False).select_related('profile')
        
        # Apply archived filter - toggle between active and archived views
        show_archived = self.request.query_params.get('show_archived', 'false').lower() == 'true'
        if show_archived:
            # Show ONLY archived accounts
            queryset = queryset.filter(profile__is_archived=True)
            logger.info(f"Showing archived users only - Found {queryset.count()} archived users")
        else:
            # Show ONLY active (non-archived) accounts
            queryset = queryset.filter(profile__is_archived=False)
            logger.info(f"Showing active users only - Found {queryset.count()} active users")
        
        # Apply position filter if provided (comma-separated positions supported)
        position = self.request.query_params.get('position', None)
        if position:
            positions = [p.strip() for p in position.split(',') if p.strip()]
            if positions:
                queryset = queryset.filter(profile__position__in=positions)
                logger.info(f"Filtering users by positions: {positions} - Found {queryset.count()} users")
        
        # Apply search filter if provided
        search = self.request.query_params.get('search', None)
        
        if search:
            queryset = (queryset.filter(
                username__icontains=search
            ) | queryset.filter(
                profile__full_name__icontains=search
            ) | queryset.filter(
                profile__email__icontains=search
            )).distinct()
            logger.info(f"Filtering users by search: '{search}' - Found {queryset.count()} users")
        
        return queryset.order_by('username')
    
    def get_serializer_class(self):
        """Use different serializers for different actions"""
        if self.action == 'create':
            return UserSerializer
        return UserListSerializer
    
    def list(self, request, *args, **kwargs):
        """Override list to add debug info to response"""
        # Log query parameters for debugging
        position = request.query_params.get('position', None)
        search = request.query_params.get('search', None)
        show_archived = request.query_params.get('show_archived', 'false')
        logger.info(f"UserViewSet.list called - position: {position}, search: {search}, show_archived: {show_archived}")
        
        response = super().list(request, *args, **kwargs)
        
        # Add debug info to response
        if hasattr(response, 'data') and isinstance(response.data, dict):
            response.data['_debug'] = {
                'position_filter': position,
                'search_filter': search,
                'show_archived': show_archived,
                'total_results': response.data.get('count', 0)
            }
        
        return response
    
    def create(self, request, *args, **kwargs):
        """Create a new user with profile"""
        # Handle both JSON and multipart form data
        data = request.data.copy()
        
        serializer = UserSerializer(data=data)
        if serializer.is_valid():
            # Save the password before hashing for email
            plain_password = request.data.get('password', '')
            username = request.data.get('username', '')
            full_name = request.data.get('full_name', '')
            email = request.data.get('email', '')
            position = request.data.get('position', '')
            
            # Create the user
            user = serializer.save()
            
            # Send welcome email with credentials
            logger.info(f"New account created for {username}, sending welcome email...")
            email_sent = send_account_created_email(
                username=username,
                password=plain_password,
                full_name=full_name,
                email=email,
                position=position
            )
            
            if email_sent:
                logger.info(f"✓ Welcome email sent to {email}")
            else:
                logger.warning(f"⚠ Failed to send welcome email to {email}")
            
            return Response({
                "message": "User created successfully",
                "user": UserListSerializer(user).data,
                "email_sent": email_sent
            }, status=status.HTTP_201_CREATED)
        return Response({
            "error": "Failed to create user",
            "details": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    def update(self, request, *args, **kwargs):
        """Update a user's details"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Handle both JSON and multipart form data
        data = request.data.copy()
        
        serializer = UserSerializer(instance, data=data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "User updated successfully",
                "user": UserListSerializer(instance).data
            }, status=status.HTTP_200_OK)
        return Response({
            "error": "Failed to update user",
            "details": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    def destroy(self, request, *args, **kwargs):
        """Permanent deletion is disabled. Use the archive action instead."""
        return Response(
            {"error": "Permanent deletion is not allowed. Use the archive endpoint instead."},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )
    
    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        """Archive a user account"""
        try:
            user = self.get_object()
            if not hasattr(user, 'profile'):
                return Response({
                    "error": "User has no profile"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if user.profile.is_archived:
                return Response({
                    "error": "User is already archived"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Archive the user
            previous_points = user.profile.points
            user.is_active = False
            user.save(update_fields=['is_active'])
            user.profile.is_archived = True
            user.profile.date_archived = timezone.now()
            user.profile.archived_by = request.user if request.user.is_authenticated else None
            user.profile.points = 0
            user.profile.save()

            log_points_change(
                entity_type='USER',
                entity_id=user.id,
                entity_name=user.profile.full_name or user.username,
                previous_points=previous_points,
                new_points=0,
                action_type='INDIVIDUAL_SET',
                changed_by=request.user if request.user.is_authenticated else None,
                reason='Account archived',
            )

            return Response({
                "message": "User archived successfully",
                "user": UserListSerializer(user).data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error archiving user {pk}: {str(e)}")
            return Response({
                "error": "Failed to archive user",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def unarchive(self, request, pk=None):
        """Unarchive/restore a user account"""
        try:
            user = self.get_object()
            if not hasattr(user, 'profile'):
                return Response({
                    "error": "User has no profile"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if not user.profile.is_archived:
                return Response({
                    "error": "User is not archived"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Unarchive the user
            user.is_active = True
            user.save(update_fields=['is_active'])
            user.profile.is_archived = False
            user.profile.date_archived = None
            user.profile.archived_by = None
            user.profile.save()
            
            return Response({
                "message": "User unarchived successfully",
                "user": UserListSerializer(user).data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error unarchiving user {pk}: {str(e)}")
            return Response({
                "error": "Failed to unarchive user",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='change_password')
    def change_password(self, request):
        """Change the logged-in user's own password (requires current password verification)"""
        if not request.user.is_authenticated:
            logger.warning("change_password: unauthenticated request")
            return Response({"error": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)

        logger.info(f"change_password called by user {request.user.username}")
        try:
            data = request.data
            current_password = data.get("current_password", "").strip()
            new_password = data.get("new_password", "").strip()
            confirm_password = data.get("confirm_password", "").strip()

            # Field presence validation
            if not current_password or not new_password or not confirm_password:
                logger.debug("change_password: missing one or more required fields")
                return Response({"error": "All fields are required"}, status=status.HTTP_400_BAD_REQUEST)

            # Password strength validation
            pw_error = validate_password_strength(new_password)
            if pw_error:
                logger.debug(f"change_password: {pw_error}")
                return Response({"error": pw_error}, status=status.HTTP_400_BAD_REQUEST)

            # Confirmation match
            if new_password != confirm_password:
                logger.debug("change_password: new passwords do not match")
                return Response({"error": "New passwords do not match"}, status=status.HTTP_400_BAD_REQUEST)

            # Verify current password
            if not request.user.check_password(current_password):
                logger.warning(f"change_password: incorrect current password for user {request.user.username}")
                return Response({"error": "Current password is incorrect"}, status=status.HTTP_400_BAD_REQUEST)

            # Apply new password and keep session alive
            from django.contrib.auth import update_session_auth_hash
            request.user.set_password(new_password)
            request.user.save()
            update_session_auth_hash(request, request.user)
            logger.info(f"✓ Password changed successfully for user {request.user.username}")

            # Send security notification email
            try:
                email = getattr(request.user, 'profile', None) and request.user.profile.email
                full_name = (getattr(request.user, 'profile', None) and request.user.profile.full_name) or request.user.username
                if email:
                    send_password_changed_email(email=email, full_name=full_name, username=request.user.username)
                    logger.info(f"✓ Password changed notification email sent to {email}")
                else:
                    logger.warning(f"change_password: no email on profile for user {request.user.username}, skipping notification")
            except Exception as email_err:
                logger.error(f"Error sending password changed notification: {str(email_err)}")
                # Don't fail the request if email fails

            return Response({"message": "Password changed successfully"}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error in change_password for user {request.user.username}: {str(e)}")
            logger.exception("Full traceback:")
            return Response({"error": "An unexpected error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='unlock_account')
    def unlock_account(self, request, pk=None):
        """Unlock a locked-out account by clearing its failed login attempts (admin action)"""
        if not request.user.is_authenticated:
            logger.warning("unlock_account: unauthenticated request")
            return Response({"error": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)

        logger.info(f"unlock_account called by {request.user.username} for pk={pk}")
        try:
            password = request.data.get("password", "").strip()
            if not password:
                logger.debug("unlock_account: no admin password provided")
                return Response({"error": "Your password is required to confirm this action"}, status=status.HTTP_400_BAD_REQUEST)

            # Verify the requesting admin's own password
            if not request.user.check_password(password):
                logger.warning(f"unlock_account: invalid admin password from {request.user.username}")
                return Response({"error": "Invalid password"}, status=status.HTTP_400_BAD_REQUEST)

            target_user = self.get_object()
            logger.info(f"unlock_account: target user is {target_user.username}")

            from .models import LoginAttempt
            if not LoginAttempt.is_locked_out(target_user.username):
                logger.info(f"unlock_account: {target_user.username} is not currently locked out")
                return Response({"error": "This account is not currently locked out"}, status=status.HTTP_400_BAD_REQUEST)

            LoginAttempt.clear_failures(target_user.username)
            logger.info(f"✓ unlock_account: cleared lockout for {target_user.username}")
            return Response({"message": f"{target_user.username} has been unlocked successfully"}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error in unlock_account for pk={pk}: {str(e)}")
            logger.exception("Full traceback:")
            return Response({"error": "An unexpected error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='send_password_reset_email')
    def send_password_reset_email(self, request, pk=None):
        """Send a password reset link email to the user (admin-initiated)"""
        try:
            user = self.get_object()
            logger.info(f"send_password_reset_email called for user pk={pk}")

            if not hasattr(user, 'profile'):
                logger.warning(f"User {pk} has no profile")
                return Response({
                    "error": "User has no profile"
                }, status=status.HTTP_400_BAD_REQUEST)

            email = user.profile.email
            if not email:
                logger.warning(f"User {pk} has no email address")
                return Response({
                    "error": "User has no email address configured"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Derive frontend URL from request origin or referer
            origin = request.META.get('HTTP_ORIGIN', '')
            referer = request.META.get('HTTP_REFERER', '')
            if origin:
                frontend_url = origin
            elif referer:
                from urllib.parse import urlparse
                parsed = urlparse(referer)
                frontend_url = f"{parsed.scheme}://{parsed.netloc}"
            else:
                frontend_url = 'http://localhost:5173'
            
            from urllib.parse import quote
            reset_url = f"{frontend_url}/password-reset?email={quote(email)}"
            logger.info(f"Constructed reset URL: {reset_url}")

            full_name = user.profile.full_name or user.username
            email_sent = send_password_reset_link_email(
                email=email,
                full_name=full_name,
                username=user.username,
                reset_url=reset_url,
            )

            if email_sent:
                logger.info(f"✓ Password reset email sent to {email} for user {user.username}")
                return Response({
                    "message": f"Password reset email sent to {email}",
                    "email_sent": True,
                }, status=status.HTTP_200_OK)
            else:
                logger.error(f"✗ Failed to send password reset email to {email}")
                return Response({
                    "error": "Failed to send email. Please try again later.",
                    "email_sent": False,
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        except Exception as e:
            logger.error(f"Error sending password reset email for user {pk}: {str(e)}")
            logger.exception("Full traceback:")
            return Response({
                "error": "Failed to send password reset email",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@method_decorator(csrf_exempt, name='dispatch')
class UnarchiveUserView(APIView):
    """Unarchive (restore) an archived user account"""
    
    def post(self, request, pk=None):
        """Restore an archived user account"""
        if not request.user.is_authenticated:
            return Response({
                "error": "Authentication required"
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            user = User.objects.select_related('profile').get(pk=pk, is_superuser=False)
        except User.DoesNotExist:
            return Response({
                "error": "User not found"
            }, status=status.HTTP_404_NOT_FOUND)
        
        if not hasattr(user, 'profile') or not user.profile.is_archived:
            return Response({
                "error": "User is not archived"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        user.is_active = True
        user.save(update_fields=['is_active'])
        user.profile.is_archived = False
        user.profile.date_archived = None
        user.profile.archived_by = None
        user.profile.save(update_fields=['is_archived', 'date_archived', 'archived_by'])
        
        return Response({
            "message": "User unarchived successfully",
            "user": UserListSerializer(user).data
        }, status=status.HTTP_200_OK)


@method_decorator(csrf_exempt, name='dispatch')
class SalesAgentsListView(APIView):
    """Lightweight endpoint to get all non-archived sales agents for dropdown selections"""
    
    def get(self, request):
        """Return all non-archived sales agents without pagination"""
        try:
            sales_agents = User.objects.filter(
                is_superuser=False,
                profile__position='Sales Agent',
                profile__is_archived=False
            ).select_related('profile').prefetch_related('team_memberships').order_by('profile__full_name')
            
            serializer = SalesAgentOptionSerializer(sales_agents, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error fetching sales agents list: {str(e)}")
            return Response({
                "error": "Failed to fetch sales agents",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@method_decorator(csrf_exempt, name='dispatch')
class CurrentUserView(APIView):
    """Get current authenticated user's profile"""
    
    def get(self, request):
        """Get current user's profile details"""
        # Manually check if user is authenticated via Django session
        if not request.user.is_authenticated:
            return Response({
                "error": "Not authenticated",
                "details": "User is not logged in. Please login first.",
                "debug": {
                    "user": str(request.user),
                    "session_key": request.session.session_key,
                    "has_sessionid_cookie": "sessionid" in request.COOKIES
                }
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            user = request.user
            serializer = UserListSerializer(user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                "error": "Failed to get user profile",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@method_decorator(csrf_exempt, name='dispatch')
class UserExportView(APIView):
    """Export users to PDF or Excel format"""
    
    # Available columns for export
    AVAILABLE_COLUMNS = {
        'id': 'ID',
        'username': 'Username',
        'full_name': 'Full Name',
        'email': 'Email',
        'position': 'Position',
        'points': 'Points',
        'status': 'Status',
    }
    
    def _get_account_status(self, user):
        """Get display status of an account"""
        if hasattr(user, 'profile'):
            if user.profile.is_activated:
                return 'Active'
        return 'Inactive'
    
    def _get_cell_value(self, user, column):
        """Get cell value for export"""
        if column == 'id':
            return user.id
        elif column == 'username':
            return user.username
        elif column == 'full_name':
            return user.profile.full_name if hasattr(user, 'profile') else ''
        elif column == 'email':
            return user.profile.email if hasattr(user, 'profile') else ''
        elif column == 'position':
            return user.profile.position if hasattr(user, 'profile') else ''
        elif column == 'points':
            return user.profile.points if hasattr(user, 'profile') else 0
        elif column == 'status':
            return self._get_account_status(user)
        return ''
    
    def _sort_users(self, users, sort_field, sort_direction):
        """Sort users by field and direction"""
        reverse = sort_direction == 'desc'
        
        if sort_field == 'status':
            return sorted(users, key=lambda u: self._get_account_status(u), reverse=reverse)
        elif sort_field in ['full_name', 'email', 'position']:
            return sorted(users, key=lambda u: getattr(u.profile, sort_field, '') if hasattr(u, 'profile') else '', reverse=reverse)
        elif sort_field == 'points':
            return sorted(users, key=lambda u: u.profile.points if hasattr(u, 'profile') else 0, reverse=reverse)
        elif sort_field == 'id':
            return sorted(users, key=lambda u: u.id, reverse=reverse)
        elif sort_field == 'username':
            return sorted(users, key=lambda u: u.username, reverse=reverse)
        return users
    
    def post(self, request):
        """Export users based on provided options"""
        # Check authentication
        if not request.user.is_authenticated:
            return Response({
                "error": "Authentication required"
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Parse request data
        columns = request.data.get('columns', list(self.AVAILABLE_COLUMNS.keys()))
        sort_field = request.data.get('sort_field', 'id')
        sort_direction = request.data.get('sort_direction', 'asc')
        export_format = request.data.get('format', 'excel')
        
        # Validate columns
        valid_columns = [c for c in columns if c in self.AVAILABLE_COLUMNS]
        if not valid_columns:
            return Response({
                "error": "No valid columns specified"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get users
        users = list(User.objects.filter(is_superuser=False).select_related('profile'))
        
        # Sort users
        users = self._sort_users(users, sort_field, sort_direction)
        
        # Generate export
        if export_format == 'pdf':
            return self._generate_pdf(users, valid_columns)
        else:
            return self._generate_excel(users, valid_columns)
    
    def _generate_excel(self, users, columns):
        """Generate Excel file"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Accounts"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            headers = [self.AVAILABLE_COLUMNS[col] for col in columns]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data rows
            for row_idx, user in enumerate(users, 2):
                for col_idx, column in enumerate(columns, 1):
                    value = self._get_cell_value(user, column)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
            
            # Adjust column widths
            column_widths = {
                'id': 8,
                'username': 20,
                'full_name': 25,
                'email': 30,
                'position': 20,
                'points': 10,
                'status': 12,
            }
            for col_idx, column in enumerate(columns, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = column_widths.get(column, 15)
            
            # Create response
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            timestamp = datetime.now().strftime('%Y-%m-%d')
            filename = f"accounts_export_{timestamp}.xlsx"
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except ImportError:
            return Response({
                "error": "Excel export not available. Please install openpyxl."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _generate_pdf(self, users, columns):
        """Generate PDF file"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            output = io.BytesIO()
            
            # Use landscape for many columns
            page_size = landscape(letter) if len(columns) > 5 else letter
            doc = SimpleDocTemplate(output, pagesize=page_size, topMargin=0.5*inch, bottomMargin=0.5*inch)
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                alignment=1,  # Center
                spaceAfter=12
            )
            elements.append(Paragraph("Accounts Export", title_style))
            
            # Subtitle with date and count
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1,
                textColor=colors.grey,
                spaceAfter=20
            )
            elements.append(Paragraph(
                f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Total Records: {len(users)}",
                subtitle_style
            ))
            
            # Table data
            headers = [self.AVAILABLE_COLUMNS[col] for col in columns]
            table_data = [headers]
            
            for user in users:
                row = [str(self._get_cell_value(user, col)) for col in columns]
                table_data.append(row)
            
            # Create table
            table = Table(table_data, repeatRows=1)
            
            # Table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ])
            table.setStyle(table_style)
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            output.seek(0)
            
            timestamp = datetime.now().strftime('%Y-%m-%d')
            filename = f"accounts_export_{timestamp}.pdf"
            
            response = HttpResponse(output.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except ImportError:
            return Response({
                "error": "PDF export not available. Please install reportlab."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@method_decorator(csrf_exempt, name='dispatch')
class BulkUpdatePointsView(APIView):
    """Bulk update points for all active accounts with password verification"""
    
    def post(self, request):
        """Apply points delta to all active accounts"""
        # Check authentication
        if not request.user.is_authenticated:
            return Response({
                "error": "Authentication required"
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Check if user is superadmin
        if not request.user.is_superuser:
            profile = getattr(request.user, 'profile', None)
            is_admin = profile and profile.position == 'Admin'
            if not is_admin:
                return Response({
                    "error": "Only superadmins can perform bulk points update"
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Get password from request
        password = request.data.get('password', '')
        if not password:
            return Response({
                "error": "Password is required for verification"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify password
        if not request.user.check_password(password):
            return Response({
                "error": "Invalid password"
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Check if this is a reset operation
        reset_to_zero = request.data.get('reset_to_zero', False)
        
        if reset_to_zero:
            points_delta = None
            operation = "reset"
        else:
            # Get points delta
            points_delta = request.data.get('points_delta')
            if points_delta is None:
                return Response({
                    "error": "Points delta is required when not resetting"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                points_delta = int(points_delta)
            except (ValueError, TypeError):
                return Response({
                    "error": "Points delta must be a valid integer"
                }, status=status.HTTP_400_BAD_REQUEST)
            operation = "update"
        
        try:
            # Get all active users (excluding superusers)
            active_users = User.objects.filter(
                is_superuser=False
            ).select_related('profile')
            
            updated_count = 0
            failed_count = 0
            failed_users = []
            
            # Generate batch_id for audit grouping
            batch_id = generate_batch_id()
            audit_entries = []
            
            # Update each user's points
            for user in active_users:
                try:
                    if hasattr(user, 'profile'):
                        old_points = user.profile.points or 0
                        if reset_to_zero:
                            user.profile.points = 0
                        else:
                            user.profile.points = old_points + points_delta
                        user.profile.save()
                        updated_count += 1
                        
                        # Collect audit entry
                        audit_entries.append({
                            'entity_type': PointsAuditLog.EntityType.USER,
                            'entity_id': user.id,
                            'entity_name': user.profile.full_name or user.username,
                            'previous_points': old_points,
                            'new_points': user.profile.points,
                            'action_type': PointsAuditLog.ActionType.BULK_RESET if reset_to_zero else PointsAuditLog.ActionType.BULK_DELTA,
                            'changed_by': request.user,
                            'reason': f'Bulk reset to 0' if reset_to_zero else f'Bulk delta {points_delta:+d}',
                            'batch_id': batch_id,
                        })
                    else:
                        failed_count += 1
                        failed_users.append(user.username)
                except Exception as e:
                    logger.error(f"Failed to update points for user {user.username}: {str(e)}")
                    failed_count += 1
                    failed_users.append(user.username)
            
            # Bulk create audit log entries
            if audit_entries:
                try:
                    bulk_log_points_changes(audit_entries)
                except Exception as e:
                    logger.error(f"Failed to create audit log entries: {str(e)}")
            
            if reset_to_zero:
                message = f"Successfully reset points to 0 for {updated_count} account(s)"
                log_message = f"Bulk points reset by {request.user.username}: Reset {updated_count} accounts to 0"
            else:
                message = f"Successfully updated points for {updated_count} account(s)"
                log_message = f"Bulk points update by {request.user.username}: {points_delta:+d} points to {updated_count} accounts"
            
            response_data = {
                "message": message,
                "updated_count": updated_count,
                "failed_count": failed_count,
                "total_affected": len(active_users),
                "operation": operation
            }
            
            if not reset_to_zero:
                response_data["points_delta"] = points_delta
            
            if failed_count > 0:
                response_data["failed_users"] = failed_users
                response_data["message"] = f"Updated {updated_count} of {len(active_users)} accounts. {failed_count} failed."
            
            logger.info(log_message)
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error in bulk points update: {str(e)}")
            return Response({
                "error": "Failed to update points",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@method_decorator(csrf_exempt, name='dispatch')
class BatchUpdatePointsView(APIView):
    """Batch update points for specific accounts (optimized single request)"""
    
    def post(self, request):
        """Update points for multiple specific accounts in a single request"""
        # Check authentication
        if not request.user.is_authenticated:
            return Response({
                "error": "Authentication required"
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        updates = request.data.get('updates', [])
        reason = request.data.get('reason', '')
        
        if not updates:
            return Response({
                "error": "No updates provided"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        updated_ids = []
        failed = []
        batch_id = generate_batch_id()
        audit_entries = []
        
        for update in updates:
            try:
                user_id = update.get('id')
                new_points = update.get('points')
                
                if user_id is None or new_points is None:
                    failed.append({'id': user_id, 'error': 'Missing id or points'})
                    continue
                
                user = User.objects.select_related('profile').get(id=user_id, is_superuser=False)
                if hasattr(user, 'profile'):
                    old_points = user.profile.points or 0
                    new_points_int = int(new_points)  # Allow negative points
                    user.profile.points = new_points_int
                    user.profile.save(update_fields=['points'])
                    updated_ids.append(user_id)
                    
                    # Collect audit entry
                    audit_entries.append({
                        'entity_type': PointsAuditLog.EntityType.USER,
                        'entity_id': user_id,
                        'entity_name': user.profile.full_name or user.username,
                        'previous_points': old_points,
                        'new_points': new_points_int,
                        'action_type': PointsAuditLog.ActionType.INDIVIDUAL_SET,
                        'changed_by': request.user,
                        'reason': reason,
                        'batch_id': batch_id,
                    })
                else:
                    failed.append({'id': user_id, 'error': 'User has no profile'})
            except User.DoesNotExist:
                failed.append({'id': user_id, 'error': 'User not found'})
            except (ValueError, TypeError) as e:
                failed.append({'id': user_id, 'error': f'Invalid points value: {str(e)}'})
            except Exception as e:
                logger.error(f"Failed to update points for user {user_id}: {str(e)}")
                failed.append({'id': user_id, 'error': str(e)})
        
        # Bulk create audit log entries
        if audit_entries:
            try:
                bulk_log_points_changes(audit_entries)
            except Exception as e:
                logger.error(f"Failed to create audit log entries: {str(e)}")
        
        return Response({
            "message": f"Updated {len(updated_ids)} account(s)",
            "updated_count": len(updated_ids),
            "failed_count": len(failed),
            "updated_ids": updated_ids,
            "failed": failed if failed else None
        }, status=status.HTTP_200_OK)
