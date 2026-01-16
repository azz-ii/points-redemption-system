import os
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from customers.models import Customer
from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'Import customers from Excel file with single column "LIST OF CUSTOMERS"'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Path to the Excel file (.xlsx)')
        parser.add_argument('--dry-run', action='store_true', help='Show what would be imported without saving')
        parser.add_argument('--skip-duplicates', action='store_true', default=True, 
                          help='Skip customers with duplicate names (default: True)')

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        dry_run = options['dry_run']
        skip_duplicates = options['skip_duplicates']

        if not os.path.exists(excel_file):
            raise CommandError(f'File {excel_file} does not exist')

        if not excel_file.endswith(('.xlsx', '.xls')):
            raise CommandError('File must be an Excel file (.xlsx or .xls)')

        self.stdout.write(f'Importing from {excel_file}...')
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN MODE - No changes will be made'))

        imported_count = 0
        skipped_count = 0
        errors = []

        try:
            # Load the workbook and get the active sheet
            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            sheet = workbook.active

            # Get existing customer names for duplicate checking
            existing_names = set(Customer.objects.values_list('name', flat=True)) if skip_duplicates else set()

            # Process rows
            header_found = False
            for row_num, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
                try:
                    # Skip empty rows
                    if not row or not any(row):
                        continue

                    # Check for header row
                    if row_num == 1:
                        # Look for "LIST OF CUSTOMERS" header (case-insensitive)
                        if row[0] and isinstance(row[0], str) and 'LIST OF CUSTOMERS' in row[0].upper():
                            header_found = True
                            self.stdout.write(f'Found header: {row[0]}')
                            continue
                        elif row[0] and isinstance(row[0], str):
                            # First row doesn't match expected header, but might be data
                            self.stdout.write(self.style.WARNING(
                                f'Expected header "LIST OF CUSTOMERS" not found. Treating first row as data.'
                            ))

                    # Get customer name from first column
                    customer_name = str(row[0]).strip() if row[0] else None

                    # Validation
                    if not customer_name:
                        errors.append(f'Row {row_num}: Empty customer name')
                        continue

                    if len(customer_name) > 255:
                        errors.append(f'Row {row_num}: Customer name too long (max 255 chars): {customer_name[:50]}...')
                        continue

                    # Check for duplicates
                    if skip_duplicates and customer_name in existing_names:
                        skipped_count += 1
                        self.stdout.write(self.style.WARNING(f'Row {row_num}: Skipping duplicate - {customer_name}'))
                        continue

                    # Create customer with placeholder values for required fields
                    if not dry_run:
                        Customer.objects.create(
                            name=customer_name,
                            contact_email=f'pending@example.com',  # Placeholder
                            phone='000-000-0000',  # Placeholder
                            location='N/A',  # Placeholder
                            points=0,
                            is_archived=False
                        )
                        existing_names.add(customer_name)  # Track within import
                        imported_count += 1
                        self.stdout.write(self.style.SUCCESS(f'Row {row_num}: Imported - {customer_name}'))
                    else:
                        # Dry run: just count
                        imported_count += 1
                        self.stdout.write(f'Row {row_num}: Would import - {customer_name}')

                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')

            workbook.close()

        except Exception as e:
            raise CommandError(f'Error reading Excel file: {str(e)}')

        # Summary
        self.stdout.write('')
        self.stdout.write('=' * 50)
        self.stdout.write(self.style.SUCCESS(f'✓ Imported: {imported_count} customers'))
        if skipped_count > 0:
            self.stdout.write(self.style.WARNING(f'⊘ Skipped (duplicates): {skipped_count}'))
        if errors:
            self.stdout.write(self.style.ERROR(f'✗ Errors: {len(errors)}'))
            self.stdout.write('')
            self.stdout.write('Error details:')
            for error in errors:
                self.stdout.write(self.style.ERROR(f'  • {error}'))
        else:
            self.stdout.write(self.style.SUCCESS('No errors'))
        self.stdout.write('=' * 50)

        if dry_run:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('NOTE: This was a dry run. Run without --dry-run to save changes.'))
            self.stdout.write(self.style.WARNING('NOTE: Placeholder values will be used for email, phone, and location.'))
