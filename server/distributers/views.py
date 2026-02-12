from django.shortcuts import render
from django.http import HttpResponse
from django.db import transaction
from django.db.models import F
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.pagination import PageNumberPagination
from rest_framework.authentication import SessionAuthentication
from rest_framework.views import APIView
from .models import Distributor
from .serializers import DistributorSerializer
import io
import logging
from datetime import datetime
from points_audit.utils import bulk_log_points_changes, generate_batch_id
from points_audit.models import PointsAuditLog

# Configure logger for distributor operations
logger = logging.getLogger('distributors')

class CsrfExemptSessionAuthentication(SessionAuthentication):
    """Session authentication without CSRF checks for API endpoints"""
    def enforce_csrf(self, request):
        return  # Skip CSRF check

class DistributorPagination(PageNumberPagination):
    """
    Pagination class for distributors list.
    """
    page_size = 15
    page_size_query_param = 'page_size'
    max_page_size = 100

class DistributorViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing distributors.
    Provides CRUD operations for distributor management.
    """
    queryset = Distributor.objects.all()
    serializer_class = DistributorSerializer
    authentication_classes = [CsrfExemptSessionAuthentication]
    permission_classes = [AllowAny]  # TEMP: Allow unauthenticated access for testing
    pagination_class = DistributorPagination
    
    def get_queryset(self):
        """
        All authenticated users can access all distributors.
        Optionally filter based on query parameters.
        By default, archived distributors are excluded unless show_archived=true.
        """
        # Start with all distributors
        queryset = Distributor.objects.all()
        
        # Filter archived distributors by default
        show_archived = self.request.query_params.get('show_archived', '').lower() == 'true'
        if not show_archived:
            queryset = queryset.filter(is_archived=False)
        
        # Apply search filter if provided
        search = self.request.query_params.get('search', None)
        
        if search:
            queryset = queryset.filter(
                name__icontains=search
            ) | queryset.filter(
                contact_email__icontains=search
            ) | queryset.filter(
                location__icontains=search
            )
        
        return queryset.order_by('name')
    
    def perform_create(self, serializer):
        """
        Set the added_by field to the current user when creating.
        """
        serializer.save(added_by=self.request.user)
    
    @action(detail=False, methods=['get'])
    def search(self, request):
        """
        Custom search endpoint for distributors.
        """
        return self.list(request)

    def destroy(self, request, *args, **kwargs):
        """Archive a distributor instead of deleting"""
        instance = self.get_object()
        instance.is_archived = True
        instance.date_archived = timezone.now()
        instance.archived_by = request.user if request.user.is_authenticated else None
        instance.save(update_fields=['is_archived', 'date_archived', 'archived_by'])
        
        return Response({
            "message": "Distributor archived successfully",
            "distributor": DistributorSerializer(instance).data
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def list_all(self, request):
        """
        Lightweight endpoint for dropdown use - returns id, name, and location only.
        No pagination for efficient single-request loading.
        """
        queryset = Distributor.objects.filter(is_archived=False).values('id', 'name', 'location').order_by('name')
        return Response(list(queryset))


@method_decorator(csrf_exempt, name='dispatch')
class DistributorExportView(APIView):
    """Export distributors to PDF or Excel format"""
    authentication_classes = [CsrfExemptSessionAuthentication]
    permission_classes = [AllowAny]
    
    # Available columns for export
    AVAILABLE_COLUMNS = {
        'id': 'ID',
        'name': 'Name',
        'contact_email': 'Contact Email',
        'phone': 'Phone',
        'location': 'Location',
        'points': 'Points',
        'date_added': 'Date Added',
        'status': 'Status',
    }
    
    def _get_cell_value(self, distributor, column):
        """Get cell value for export"""
        if column == 'id':
            return distributor.id
        elif column == 'name':
            return distributor.name
        elif column == 'contact_email':
            return distributor.contact_email or ''
        elif column == 'phone':
            return distributor.phone or ''
        elif column == 'location':
            return distributor.location or ''
        elif column == 'points':
            return distributor.points or 0
        elif column == 'date_added':
            return distributor.date_added.strftime('%Y-%m-%d') if distributor.date_added else ''
        elif column == 'status':
            return 'Archived' if distributor.is_archived else 'Active'
        return ''
    
    def _sort_distributors(self, distributors, sort_field, sort_direction):
        """Sort distributors by field and direction"""
        reverse = sort_direction == 'desc'
        
        if sort_field == 'date_added':
            return sorted(distributors, key=lambda d: d.date_added or datetime.min.date(), reverse=reverse)
        elif sort_field in ['id', 'points']:
            return sorted(distributors, key=lambda d: getattr(d, sort_field, 0), reverse=reverse)
        elif sort_field in ['name', 'contact_email', 'phone', 'location']:
            return sorted(distributors, key=lambda d: getattr(d, sort_field, '').lower(), reverse=reverse)
        return distributors
    
    def post(self, request):
        """Export distributors based on provided options"""
        # Parse request data
        columns = request.data.get('columns', list(self.AVAILABLE_COLUMNS.keys()))
        sort_field = request.data.get('sort_field', 'id')
        sort_direction = request.data.get('sort_direction', 'asc')
        export_format = request.data.get('format', 'excel')
        
        # Validate columns
        valid_columns = [c for c in columns if c in self.AVAILABLE_COLUMNS]
        if not valid_columns:
            return Response({
                "error": "No valid columns specified"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get distributors (exclude archived)
        distributors = list(Distributor.objects.filter(is_archived=False))
        
        # Sort distributors
        distributors = self._sort_distributors(distributors, sort_field, sort_direction)
        
        # Generate export
        if export_format == 'pdf':
            return self._generate_pdf(distributors, valid_columns)
        else:
            return self._generate_excel(distributors, valid_columns)
    
    def _generate_excel(self, distributors, columns):
        """Generate Excel file"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Distributors"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            headers = [self.AVAILABLE_COLUMNS[col] for col in columns]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data rows
            for row_idx, distributor in enumerate(distributors, 2):
                for col_idx, column in enumerate(columns, 1):
                    value = self._get_cell_value(distributor, column)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
            
            # Adjust column widths
            column_widths = {
                'id': 8,
                'name': 25,
                'contact_email': 30,
                'phone': 15,
                'location': 25,
                'points': 10,
                'date_added': 15,
                'status': 12,
            }
            for col_idx, column in enumerate(columns, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = column_widths.get(column, 15)
            
            # Create response
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            timestamp = datetime.now().strftime('%Y-%m-%d')
            filename = f"distributors_export_{timestamp}.xlsx"
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except ImportError:
            return Response({
                "error": "Excel export not available. Please install openpyxl."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _generate_pdf(self, distributors, columns):
        """Generate PDF file"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            output = io.BytesIO()
            
            # Use landscape for many columns
            page_size = landscape(letter) if len(columns) > 5 else letter
            doc = SimpleDocTemplate(output, pagesize=page_size, topMargin=0.5*inch, bottomMargin=0.5*inch)
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                alignment=1,  # Center
                spaceAfter=12
            )
            elements.append(Paragraph("Distributors Export", title_style))
            
            # Subtitle with date and count
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1,
                textColor=colors.grey,
                spaceAfter=20
            )
            elements.append(Paragraph(
                f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Total Records: {len(distributors)}",
                subtitle_style
            ))
            
            # Table data
            headers = [self.AVAILABLE_COLUMNS[col] for col in columns]
            table_data = [headers]
            
            for distributor in distributors:
                row = [str(self._get_cell_value(distributor, col)) for col in columns]
                table_data.append(row)
            
            # Create table
            table = Table(table_data, repeatRows=1)
            
            # Table style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ])
            table.setStyle(table_style)
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            output.seek(0)
            
            timestamp = datetime.now().strftime('%Y-%m-%d')
            filename = f"distributors_export_{timestamp}.pdf"
            
            response = HttpResponse(output.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except ImportError:
            return Response({
                "error": "PDF export not available. Please install reportlab."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@method_decorator(csrf_exempt, name='dispatch')
class DistributorBulkUpdatePointsView(APIView):
    """Bulk update points for all distributors with password verification (optimized with single query)"""
    authentication_classes = [CsrfExemptSessionAuthentication]
    permission_classes = []
    
    def post(self, request):
        """Apply points delta to all distributors using optimized bulk operations"""
        # Check authentication
        if not request.user.is_authenticated:
            return Response({
                "error": "Authentication required"
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Check if user is superadmin or admin
        if not request.user.is_superuser:
            profile = getattr(request.user, 'profile', None)
            is_admin = profile and profile.position == 'Admin'
            if not is_admin:
                return Response({
                    "error": "Only superadmins can perform bulk points update"
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Get password from request
        password = request.data.get('password', '')
        if not password:
            return Response({
                "error": "Password is required for verification"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify password
        if not request.user.check_password(password):
            return Response({
                "error": "Invalid password"
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Check if this is a reset operation
        reset_to_zero = request.data.get('reset_to_zero', False)
        
        if reset_to_zero:
            points_delta = None
            operation = "reset"
        else:
            # Get points delta
            points_delta = request.data.get('points_delta')
            if points_delta is None:
                return Response({
                    "error": "Points delta is required when not resetting"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                points_delta = int(points_delta)
            except (ValueError, TypeError):
                return Response({
                    "error": "Points delta must be a valid integer"
                }, status=status.HTTP_400_BAD_REQUEST)
            operation = "update"
        
        try:
            # Use atomic transaction for data consistency
            with transaction.atomic():
                # Snapshot current points for audit logging (exclude archived)
                all_distributors = list(Distributor.objects.filter(is_archived=False).values('id', 'name', 'points'))
                
                if reset_to_zero:
                    # Single SQL UPDATE query to reset all non-archived distributors to 0
                    updated_count = Distributor.objects.filter(is_archived=False).update(points=0)
                    message = f"Successfully reset points to 0 for {updated_count} distributor(s)"
                    log_message = f"Bulk points reset by {request.user.username}: Reset {updated_count} distributors to 0"
                else:
                    # Single SQL UPDATE query using F() expression for atomic increment (exclude archived)
                    updated_count = Distributor.objects.filter(is_archived=False).update(points=F('points') + points_delta)
                    message = f"Successfully updated points for {updated_count} distributor(s)"
                    log_message = f"Bulk points update by {request.user.username}: {points_delta:+d} points to {updated_count} distributors"
            
            # Create audit log entries from snapshot
            batch_id = generate_batch_id()
            audit_entries = []
            for d in all_distributors:
                old_pts = d['points'] or 0
                new_pts = 0 if reset_to_zero else old_pts + points_delta
                audit_entries.append({
                    'entity_type': PointsAuditLog.EntityType.DISTRIBUTOR,
                    'entity_id': d['id'],
                    'entity_name': d['name'],
                    'previous_points': old_pts,
                    'new_points': new_pts,
                    'action_type': PointsAuditLog.ActionType.BULK_RESET if reset_to_zero else PointsAuditLog.ActionType.BULK_DELTA,
                    'changed_by': request.user,
                    'reason': 'Bulk reset to 0' if reset_to_zero else f'Bulk delta {points_delta:+d}',
                    'batch_id': batch_id,
                })
            if audit_entries:
                try:
                    bulk_log_points_changes(audit_entries)
                except Exception as e:
                    logger.error(f"Failed to create audit log entries: {str(e)}")
            
            response_data = {
                "message": message,
                "updated_count": updated_count,
                "failed_count": 0,
                "total_affected": updated_count,
                "operation": operation
            }
            
            if not reset_to_zero:
                response_data["points_delta"] = points_delta
            
            logger.info(log_message)
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error in bulk points update: {str(e)}")
            return Response({
                "error": "Failed to update points",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@method_decorator(csrf_exempt, name='dispatch')
class DistributorBatchUpdatePointsView(APIView):
    """Batch update points for specific distributors (optimized with bulk_update)"""
    authentication_classes = [CsrfExemptSessionAuthentication]
    permission_classes = []
    
    def post(self, request):
        """Update points for multiple specific distributors using bulk_update for performance"""
        # Check authentication
        if not request.user.is_authenticated:
            return Response({
                "error": "Authentication required"
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        updates = request.data.get('updates', [])
        reason = request.data.get('reason', '')
        
        if not updates:
            return Response({
                "error": "No updates provided"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate and prepare updates
        update_map = {}
        failed = []
        
        for update in updates:
            try:
                distributor_id = update.get('id')
                new_points = update.get('points')
                
                if distributor_id is None or new_points is None:
                    failed.append({'id': distributor_id, 'error': 'Missing id or points'})
                    continue
                
                # Convert and validate points value
                try:
                    new_points = int(new_points)
                except (ValueError, TypeError) as e:
                    failed.append({'id': distributor_id, 'error': f'Invalid points value: {str(e)}'})
                    continue
                
                update_map[distributor_id] = new_points
                
            except Exception as e:
                logger.error(f"Error preparing update for distributor {distributor_id}: {str(e)}")
                failed.append({'id': distributor_id, 'error': str(e)})
        
        # Fetch all distributors that need updating (exclude archived)
        distributor_ids = list(update_map.keys())
        distributors_to_update = Distributor.objects.filter(id__in=distributor_ids, is_archived=False)
        
        # Check for missing distributors (not found or archived)
        found_ids = set(distributors_to_update.values_list('id', flat=True))
        missing_ids = set(distributor_ids) - found_ids
        for missing_id in missing_ids:
            failed.append({'id': missing_id, 'error': 'Distributor not found or archived'})
        
        # Snapshot old points and apply updates to distributor objects
        old_points_map = {}
        updated_distributors = []
        for distributor in distributors_to_update:
            old_points_map[distributor.id] = distributor.points or 0
            distributor.points = update_map[distributor.id]
            updated_distributors.append(distributor)
        
        # Perform bulk update in a transaction
        updated_ids = []
        try:
            with transaction.atomic():
                if updated_distributors:
                    Distributor.objects.bulk_update(updated_distributors, ['points'])
                    updated_ids = [d.id for d in updated_distributors]
        except Exception as e:
            logger.error(f"Failed to bulk update distributors: {str(e)}")
            return Response({
                "error": "Failed to update distributors",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Create audit log entries
        if updated_ids:
            batch_id = generate_batch_id()
            audit_entries = []
            for d in updated_distributors:
                if d.id in updated_ids:
                    audit_entries.append({
                        'entity_type': PointsAuditLog.EntityType.DISTRIBUTOR,
                        'entity_id': d.id,
                        'entity_name': d.name,
                        'previous_points': old_points_map.get(d.id, 0),
                        'new_points': d.points,
                        'action_type': PointsAuditLog.ActionType.INDIVIDUAL_SET,
                        'changed_by': request.user,
                        'reason': reason,
                        'batch_id': batch_id,
                    })
            if audit_entries:
                try:
                    bulk_log_points_changes(audit_entries)
                except Exception as e:
                    logger.error(f"Failed to create audit log entries: {str(e)}")
        
        return Response({
            "message": f"Updated {len(updated_ids)} distributor(s)",
            "updated_count": len(updated_ids),
            "failed_count": len(failed),
            "updated_ids": updated_ids,
            "failed": failed if failed else None
        }, status=status.HTTP_200_OK)


@method_decorator(csrf_exempt, name='dispatch')
class UnarchiveDistributorView(APIView):
    """Unarchive (restore) an archived distributor"""
    authentication_classes = [CsrfExemptSessionAuthentication]
    permission_classes = [AllowAny]
    
    def post(self, request, pk=None):
        """Restore an archived distributor"""
        try:
            distributor = Distributor.objects.get(pk=pk)
        except Distributor.DoesNotExist:
            return Response({"error": "Distributor not found"}, status=status.HTTP_404_NOT_FOUND)
        
        if not distributor.is_archived:
            return Response({"error": "Distributor is not archived"}, status=status.HTTP_400_BAD_REQUEST)
        
        distributor.is_archived = False
        distributor.date_archived = None
        distributor.archived_by = None
        distributor.save(update_fields=['is_archived', 'date_archived', 'archived_by'])
        
        return Response({
            "message": "Distributor unarchived successfully",
            "distributor": DistributorSerializer(distributor).data
        }, status=status.HTTP_200_OK)
